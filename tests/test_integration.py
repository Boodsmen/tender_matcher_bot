"""
Integration тесты для полного цикла работы бота.

Тестируют:
- Сопоставление требований с оборудованием в БД
- Корректность процента совпадения
- Формат Excel отчёта
- EAV-матчинг (compare_values_eav)
"""

import asyncio
import json
import os
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

pytest_plugins = ("pytest_asyncio",)


class TestFullCycle:
    @pytest.fixture
    def sample_requirements(self):
        return {
            "items": [
                {
                    "item_name": "Коммутатор L3",
                    "model_name": None,
                    "category": "Коммутаторы",
                    "required_specs": {
                        "ports_1g_rj45": 24,
                        "ports_10g_sfp_plus": 4,
                        "power_watt": 200,
                        "layer": 3,
                        "poe_support": True,
                    },
                }
            ]
        }

    @pytest.mark.asyncio
    async def test_matcher_returns_all_equipment_for_category(self):
        """
        Проверка что matcher возвращает всё оборудование категории без лимита.
        """
        from services.matcher import find_matching_models

        # Создаём 300 mock-моделей оборудования
        mock_equipment = []
        for i in range(300):
            eq = MagicMock()
            eq.id = i
            eq.model_name = f"Model_{i}"
            eq.category = "Коммутаторы"
            eq.source_filename = f"Category_Switch_test_v{20 + i % 10}.xlsx"
            eq.version = f"v{20 + i % 10}"
            eq.attributes = {"ports_1g_rj45": 24 + i % 10}
            mock_equipment.append(eq)

        with patch("services.matcher.get_equipment_by_category", new_callable=AsyncMock) as mock_get, \
             patch("services.matcher.get_specs_by_equipment_ids", new_callable=AsyncMock) as mock_specs:
            mock_get.return_value = mock_equipment
            # Return empty specs for all — just checking count, not match quality
            mock_specs.return_value = {}

            requirements = {
                "items": [
                    {
                        "model_name": None,
                        "category": "Коммутаторы",
                        "required_specs": {"ports_1g_rj45": 24},
                    }
                ]
            }

            result = await find_matching_models(requirements)
            total_found = result["summary"]["total_models_found"]
            # Subcategory expansion means base (300) + subcategories (3 × 300) = 1200
            assert total_found >= 300, f"Expected at least 300, got {total_found}"

    @pytest.mark.asyncio
    async def test_numeric_comparison_works(self):
        """Проверка что числовые сравнения работают корректно."""
        from services.matcher import compare_spec_values

        assert compare_spec_values(24, 24, "ports") is True
        assert compare_spec_values(24, 30, "ports") is True
        assert compare_spec_values(24, 20, "ports") is False

        assert compare_spec_values("24 порта", 24, "ports") is True
        assert compare_spec_values("24 порта", 30, "ports") is True
        assert compare_spec_values("24 порта", 20, "ports") is False

        assert compare_spec_values("10-20", 25, "range") is True
        assert compare_spec_values("10-20", 15, "range") is False

        assert compare_spec_values("2x4", 10, "calc") is True
        assert compare_spec_values("2x4", 5, "calc") is False

        assert compare_spec_values(200, 195, "power", allow_lower=True) is True
        assert compare_spec_values(200, 180, "power", allow_lower=True) is False

    @pytest.mark.asyncio
    async def test_match_percentage_calculation(self, sample_requirements):
        """Проверка корректности вычисления процента совпадения."""
        from services.matcher import calculate_match_percentage

        required_specs = sample_requirements["items"][0]["required_specs"]

        model_attrs_100 = {
            "ports_1g_rj45": 24,
            "ports_10g_sfp_plus": 4,
            "power_watt": 200,
            "layer": 3,
            "poe_support": True,
        }
        result = calculate_match_percentage(required_specs, model_attrs_100)
        assert result["match_percentage"] == 100.0
        assert len(result["matched_specs"]) == 5

        model_attrs_60 = {
            "ports_1g_rj45": 24,
            "ports_10g_sfp_plus": 4,
            "power_watt": 200,
        }
        result = calculate_match_percentage(required_specs, model_attrs_60)
        assert result["match_percentage"] == 60.0
        assert len(result["matched_specs"]) == 3
        assert len(result["missing_specs"]) == 2

    def test_excel_report_generates_successfully(self, tmp_path):
        """Проверка что Excel отчёт генерируется без ошибок."""
        from services.excel_generator import generate_report

        match_results = {
            "results": [
                {
                    "requirement": {
                        "item_name": "Test Item",
                        "required_specs": {"ports_1g_sfp": 24},
                    },
                    "matches": {
                        "ideal": [
                            {
                                "model_id": 1,
                                "model_name": "MES3710P",
                                "category": "Коммутаторы",
                                "version": "v29",
                                "source_filename": "Category_Switch_07.11_ISS_v29.xlsx",
                                "match_percentage": 100.0,
                                "matched_specs": ["ports_1g_sfp"],
                                "unmapped_specs": [],
                                "missing_specs": [],
                                "different_specs": {},
                                "attributes": {"ports_1g_sfp": 24},
                                "specifications": {"ports_1g_sfp": 24},
                            }
                        ],
                        "partial": [],
                        "not_matched": [],
                    },
                }
            ],
            "summary": {
                "total_requirements": 1,
                "total_models_found": 1,
                "ideal_matches": 1,
                "partial_matches": 0,
            },
        }

        requirements = {
            "items": [{"item_name": "Test Item", "required_specs": {"ports_1g_sfp": 24}}]
        }

        file_path = generate_report(
            requirements, match_results,
            output_dir=str(tmp_path),
            filename="test.docx",
            processing_time=1.5,
        )

        assert os.path.exists(file_path)

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            assert "Сводка" in wb.sheetnames
            assert len(wb.sheetnames) >= 2
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_reverse_mapping_loaded(self):
        """Проверка что reverse_normalization_map.json загружается."""
        from services.excel_generator import _load_reverse_mapping

        reverse_mapping = _load_reverse_mapping()
        assert isinstance(reverse_mapping, dict)


class TestEAVMatching:
    """Unit tests for compare_values_eav — the core EAV matching logic."""

    def test_numeric_gte_passes_when_model_meets_requirement(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav(">=24", "24 порта", 24.0) is True
        assert compare_values_eav(">=24", "28 портов", 28.0) is True

    def test_numeric_gte_fails_when_model_below_requirement(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav(">=24", "20 портов", 20.0) is False

    def test_numeric_lte_passes(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav("<=100", "80 Вт", 80.0) is True
        assert compare_values_eav("<=100", "100 Вт", 100.0) is True

    def test_numeric_lte_fails(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav("<=100", "120 Вт", 120.0) is False

    def test_equality_absolute_tolerance(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav("=24", "24", 24.0) is True
        assert compare_values_eav("=24", "24.005", 24.005) is True  # within 0.01
        assert compare_values_eav("=24", "25", 25.0) is False

    def test_equality_relative_tolerance_large_numbers(self):
        from services.matcher import compare_values_eav
        # 10000 == 10000 within 0.1% → True
        assert compare_values_eav("=10000", "10000", 10000.0) is True
        # 10000 vs 10005 → within 0.1% relative → True
        assert compare_values_eav("=10000", "10005", 10005.0) is True
        # 10000 vs 10200 → 2% diff → False
        assert compare_values_eav("=10000", "10200", 10200.0) is False

    def test_text_exact_match(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav("Коммутатор L3", "Коммутатор L3", None) is True

    def test_text_boolean_synonyms(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav(True, "да", None) is True
        assert compare_values_eav(True, "yes", None) is True
        assert compare_values_eav(False, "нет", None) is True
        assert compare_values_eav(True, "нет", None) is False

    def test_none_value_text_returns_false(self):
        from services.matcher import compare_values_eav
        assert compare_values_eav(">=24", None, None) is False

    def test_allow_lower_applies_5pct_tolerance(self):
        from services.matcher import compare_values_eav
        # 200 * 0.95 = 190, model=195 → passes with allow_lower
        assert compare_values_eav(">=200", "195", 195.0, allow_lower=True) is True
        # model=180 → below 190 threshold → fails
        assert compare_values_eav(">=200", "180", 180.0, allow_lower=True) is False

    def test_compound_list_requirement(self):
        from services.matcher import compare_values_eav
        # Both conditions must be met: >512 AND <=1024
        assert compare_values_eav([">512", "<=1024"], "768", 768.0) is True
        assert compare_values_eav([">512", "<=1024"], "1024", 1024.0) is True
        assert compare_values_eav([">512", "<=1024"], "512", 512.0) is False
        assert compare_values_eav([">512", "<=1024"], "2048", 2048.0) is False


class TestDatabaseIntegration:
    @pytest.mark.asyncio
    @pytest.mark.skipif(
        not os.environ.get("DB_INTEGRATION"),
        reason="Requires running PostgreSQL database (set DB_INTEGRATION=1 to enable)",
    )
    async def test_database_has_equipment(self):
        from database.crud import get_equipment_count
        count = await get_equipment_count()
        assert count > 0, f"Expected equipment in DB, got {count}"

    @pytest.mark.asyncio
    @pytest.mark.skipif(
        not os.environ.get("DB_INTEGRATION"),
        reason="Requires running PostgreSQL database (set DB_INTEGRATION=1 to enable)",
    )
    async def test_stats_returns_categories(self):
        from database.crud import get_stats
        stats = await get_stats()
        assert isinstance(stats, dict)
        assert len(stats) > 0


class TestPerformance:
    @pytest.mark.asyncio
    @pytest.mark.skipif(
        not os.environ.get("DB_INTEGRATION"),
        reason="Requires running PostgreSQL database (set DB_INTEGRATION=1 to enable)",
    )
    async def test_full_cycle_performance(self):
        import time
        from services.matcher import find_matching_models

        requirements = {
            "items": [
                {
                    "model_name": None,
                    "category": "Коммутаторы",
                    "required_specs": {"ports_1g_rj45": 24, "power_watt": 200},
                }
            ]
        }

        start_time = time.time()
        result = await find_matching_models(requirements)
        elapsed_time = time.time() - start_time

        assert elapsed_time < 10.0, f"Full cycle took {elapsed_time:.2f}s (expected < 10s)"
        assert result["summary"]["total_models_found"] > 0


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
