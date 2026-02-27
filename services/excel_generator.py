"""
Генерация Excel отчётов с результатами сопоставления моделей.

Структура отчёта:
- Лист "Сводка"         — плашка предупреждения + метаданные +
                          по одной таблице на каждую позицию ТЗ (рядом или стопкой)
- Листы "Поз. N ..."    — один лист на каждую позицию ТЗ:
                          таблица «Характеристика | Требуется | Факт. (модель 1) | ...»
"""

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from utils.logger import logger

# Пороги цветовой схемы
THRESHOLD_GREEN  = 85.0   # ≥ 85% → зелёный
THRESHOLD_YELLOW = 70.0   # 70–84% → жёлтый
THRESHOLD_ORANGE = 50.0   # 50–69% → оранжевый
THRESHOLD_MIN    = 50.0   # < 50% → не показываем

# Цвета ячеек
COLOR_GREEN       = "C6EFCE"
COLOR_YELLOW      = "FFEB9C"
COLOR_ORANGE      = "FFD699"
COLOR_RED         = "FFC7CE"
COLOR_GRAY        = "D9D9D9"
COLOR_LIGHT_GRAY  = "F2F2F2"
COLOR_BLUE_HDR    = "BDD7EE"
COLOR_WARN_BG     = "FF0000"
COLOR_WARN_TEXT   = "FFFFFF"

_REVERSE_MAPPING_CACHE: Optional[Dict[str, str]] = None


def _load_reverse_mapping() -> Dict[str, str]:
    global _REVERSE_MAPPING_CACHE
    if _REVERSE_MAPPING_CACHE is not None:
        return _REVERSE_MAPPING_CACHE
    try:
        path = Path(__file__).parent.parent / "data" / "reverse_normalization_map.json"
        with open(path, "r", encoding="utf-8") as f:
            _REVERSE_MAPPING_CACHE = json.load(f)
    except Exception as e:
        logger.warning(f"Failed to load reverse_normalization_map.json: {e}")
        _REVERSE_MAPPING_CACHE = {}
    return _REVERSE_MAPPING_CACHE


def _readable_key(key: str) -> str:
    return _load_reverse_mapping().get(key, key.replace("_", " ").title())



def _parse_version(source_file: str) -> str:
    if not source_file:
        return "—"
    m = re.search(r'finalUPDv\.(\d+)\.(\d+)', source_file)
    if m:
        return f"finalUPD v{m.group(1)}.{m.group(2)}"
    if 'finalUPD' in source_file:
        return "finalUPD"
    m = re.search(r'v(\d+)(?:\.(\d+))?', source_file)
    if m:
        v = f"v{m.group(1)}"
        if m.group(2):
            v += f".{m.group(2)}"
        if '_new' in source_file:
            v += " (new)"
        return v
    return source_file



def _pct_color(pct: float) -> str:
    if pct >= THRESHOLD_GREEN:
        return COLOR_GREEN
    if pct >= THRESHOLD_YELLOW:
        return COLOR_YELLOW
    if pct >= THRESHOLD_ORANGE:
        return COLOR_ORANGE
    return COLOR_RED



def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _bold(size: int = 11, color: str = None) -> Font:
    kwargs = {"bold": True, "size": size}
    if color:
        kwargs["color"] = color
    return Font(**kwargs)


def _center(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_width(ws, min_w: int = 8, max_w: int = 60) -> None:
    for col in ws.columns:
        best = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    best = max(best, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(best + 2, min_w), max_w)


def _set_row(ws, row: int, values: list, fills=None, fonts=None, aligns=None,
             height: int = None, border: bool = False) -> None:
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        if fills and ci - 1 < len(fills) and fills[ci - 1]:
            cell.fill = _fill(fills[ci - 1])
        if fonts and ci - 1 < len(fonts) and fonts[ci - 1]:
            cell.font = fonts[ci - 1]
        if aligns and ci - 1 < len(aligns) and aligns[ci - 1]:
            cell.alignment = aligns[ci - 1]
        if border:
            cell.border = _thin_border()
    if height:
        ws.row_dimensions[row].height = height


def _merge_row(ws, row: int, start_col: int, end_col: int, value,
               fill_color: str = None, font=None, align=None, height: int = None) -> None:
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    if fill_color:
        cell.fill = _fill(fill_color)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if height:
        ws.row_dimensions[row].height = height


def _comparison_detail(req_val: Any, mod_val: Any) -> str:
    try:
        from services.matcher import extract_number, extract_number_with_operator
        req_num, op = extract_number_with_operator(req_val)
        mod_num = extract_number(mod_val)
        if req_num is not None and mod_num is not None:
            return f"{mod_num} {op} {req_num}"
    except Exception:
        pass
    return ""



def _create_summary_sheet(
    wb: Workbook,
    match_results: Dict[str, Any],
    filename: str,
    processing_time: float,
) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_properties.tabColor = "2E75B6"
    ws.column_dimensions["A"].width = 2

    _nw_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _nw_left   = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)

    results = match_results.get("results", [])
    n_items = len(results)

    n_cols = 10
    _merge_row(ws, 1, 1, n_cols,
               "Отчёт по подбору оборудования",
               fill_color="2E75B6",
               font=Font(bold=True, size=16, color="FFFFFF"),
               align=_nw_center, height=36)

    _merge_row(ws, 2, 1, n_cols,
               "Программа может допускать ошибки — проверьте модели по исходным таблицам каталога",
               fill_color="FF8C00",
               font=Font(bold=True, size=10, color="FFFFFF"),
               align=_nw_center, height=22)

    ws.row_dimensions[3].height = 6

    total_specs_all = sum(
        len(_effective_specs(r.get("requirement", {}).get("required_specs", {})))
        for r in results
    )
    meta_row = 4
    meta_items = [
        ("Файл ТЗ", filename),
        ("Дата", datetime.now().strftime("%d.%m.%Y  %H:%M")),
        ("Время обработки", f"{processing_time:.1f} сек" if processing_time else "—"),
        ("Позиций", str(n_items)),
        ("Характеристик", str(total_specs_all)),
        ("Порог", f"≥ {THRESHOLD_MIN:.0f}%"),
    ]
    for label, value in meta_items:
        lc = ws.cell(row=meta_row, column=2, value=label)
        lc.font = Font(size=10, color="666666")
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws.cell(row=meta_row, column=3, value=value)
        vc.font = Font(size=10, bold=True)
        vc.alignment = _nw_left
        ws.merge_cells(start_row=meta_row, start_column=3,
                       end_row=meta_row, end_column=5)
        ws.row_dimensions[meta_row].height = 18
        meta_row += 1

    no_category_items = [
        r["requirement"].get("item_name") or r["requirement"].get("model_name") or f"Позиция {i+1}"
        for i, r in enumerate(results)
        if r.get("category_not_detected")
    ]
    if no_category_items:
        names_str = ", ".join(no_category_items[:5])
        if len(no_category_items) > 5:
            names_str += f" и ещё {len(no_category_items) - 5}"
        cat_warn_cell = ws.cell(
            row=meta_row, column=2,
            value=f"Категория не определена ({len(no_category_items)} поз.): {names_str}"
        )
        cat_warn_cell.font = Font(bold=True, size=10, color="C00000")
        ws.merge_cells(start_row=meta_row, start_column=2, end_row=meta_row, end_column=9)
        ws.row_dimensions[meta_row].height = 20
        meta_row += 1

    # ── Разделитель перед таблицами ───────────────────────────────────────────
    ws.row_dimensions[meta_row].height = 12
    meta_row += 1

    cur = meta_row
    for idx, result in enumerate(results, 1):
        cur = _summary_single_item_block(ws, result, idx, start_row=cur, col_offset=1)
        cur += 1


def _summary_single_item_block(ws, result: dict, idx: int,
                                start_row: int, col_offset: int) -> int:
    """Рисует блок одной позиции. Возвращает следующую свободную строку."""
    _nw_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _nw_left   = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)

    req = result["requirement"]
    name = req.get("item_name") or req.get("model_name") or f"Позиция {idx}"
    n_specs = len(_effective_specs(req.get("required_specs", {})))

    _merge_row(ws, start_row, col_offset + 1, col_offset + 7,
               f"Позиция {idx}: {name}   ({n_specs} хар-к)",
               fill_color="2E75B6",
               font=Font(bold=True, size=11, color="FFFFFF"),
               align=_nw_center, height=28)

    hdr_row = start_row + 1
    hdrs = [
        ("№",          4),
        ("Модель",    26),
        ("Версия",    14),
        ("Совпадение", 13),
        ("Совпало",    10),
        ("Не совп.",   10),
        ("Нет данных", 11),
    ]
    for ci, (h, w) in enumerate(hdrs, col_offset + 1):
        cell = ws.cell(row=hdr_row, column=ci, value=h)
        cell.font = Font(bold=True, size=9, color="333333")
        cell.fill = _fill(COLOR_BLUE_HDR)
        cell.alignment = _nw_center
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[hdr_row].height = 20

    # Данные
    data_row = hdr_row + 1
    last_row = hdr_row
    top_models = _collect_top_models(result, limit=15, filter_by_spec_count=_FILTER_BY_SPEC_COUNT)

    ZEBRA_A = "F7F9FC"
    ZEBRA_B = "FFFFFF"

    if not top_models:
        cell = ws.cell(row=data_row, column=col_offset + 1,
                       value="Нет моделей с совпадением ≥70% и достаточным числом характеристик")
        cell.font = Font(italic=True, size=10, color="999999")
        cell.alignment = _nw_left
        ws.merge_cells(start_row=data_row, start_column=col_offset + 1,
                       end_row=data_row, end_column=col_offset + len(hdrs))
        ws.row_dimensions[data_row].height = 22
        last_row = data_row

    for ri, item in enumerate(top_models, 0):
        pct = item["pct"]
        bg = _pct_color(pct)
        zebra = ZEBRA_A if ri % 2 == 0 else ZEBRA_B
        vals = [ri + 1, item["model_name"], item["version"],
                f"{pct:.0f}%", item["matched"],
                item["different"], item["unmapped"]]
        for ci, v in enumerate(vals, col_offset + 1):
            cell = ws.cell(row=data_row + ri, column=ci, value=v)
            if ci == col_offset + 4:  # колонка "Совпадение" — цветная
                cell.fill = _fill(bg)
                cell.font = Font(bold=True, size=10)
            elif ci == col_offset + 2:
                cell.fill = _fill(zebra)
                cell.font = Font(bold=True, size=10)
            else:
                cell.fill = _fill(zebra)
                cell.font = Font(size=10)
            cell.alignment = _nw_center if ci != col_offset + 2 else _nw_left
            cell.border = _thin_border()
        ws.row_dimensions[data_row + ri].height = 20
        last_row = data_row + ri

    return last_row + 1


def _collect_top_models(result: dict, limit: int = 15,
                        filter_by_spec_count: bool = True) -> list:
    """
    Собирает топ-N моделей для позиции, отсортированных по % совпадения.

    Фильтры:
    - < THRESHOLD_MIN% совпадения → не показываем
    - filter_by_spec_count=True: модели с меньшим числом характеристик чем в ТЗ → не показываем
    """
    req = result["requirement"]
    total_specs = len(_effective_specs(req.get("required_specs", {})))
    rows = []
    for cat in ("ideal", "partial", "not_matched"):
        for m in result["matches"].get(cat, []):
            pct = m["match_percentage"]
            if pct < THRESHOLD_MIN:
                continue
            model_specs = m.get("attributes") or m.get("specifications") or {}
            if filter_by_spec_count and len(model_specs) < total_specs:
                continue
            rows.append({
                "model_name": m["model_name"],
                "version": m.get("version") or _parse_version(m.get("source_filename", m.get("source_file", ""))),
                "pct": pct,
                "matched": len(m.get("matched_specs", [])),
                "different": len(m.get("different_specs", {})),
                "unmapped": len(m.get("unmapped_specs", m.get("missing_specs", []))),
                "total_specs": total_specs,
                "model_total_specs": len(model_specs),
                "match_data": m,
            })
    rows.sort(key=lambda x: x["pct"], reverse=True)
    return rows[:limit]



def _create_detail_sheet(
    wb: Workbook,
    result: dict,
    position_idx: int,
    max_models: int = 10,
) -> None:
    """
    Один лист для одной позиции ТЗ.
    Колонки: № | Характеристика | Требуется | Модель1 | Модель2 | ...
    """
    req = result["requirement"]
    item_name = req.get("item_name") or req.get("model_name") or f"Позиция {position_idx}"
    required_specs: Dict[str, Any] = _effective_specs(req.get("required_specs", {}))
    reverse_mapping = _load_reverse_mapping()

    top = _collect_top_models(result, limit=max_models, filter_by_spec_count=_FILTER_BY_SPEC_COUNT)
    if not top:
        top = []

    safe_name = re.sub(r'[\[\]:*?/\\]', '', item_name)[:28]
    sheet_name = f"Поз.{position_idx} {safe_name}".strip()
    existing_titles = {s.title for s in wb.worksheets}
    if sheet_name in existing_titles:
        sheet_name = f"Поз.{position_idx}_{position_idx}"

    ws = wb.create_sheet(sheet_name)

    n_models = len(top)
    COL_NUM    = 1
    COL_CHAR   = 2
    COL_REQ    = 3
    COL_FIRST  = 4
    total_cols = COL_FIRST + n_models - 1 if n_models else COL_FIRST

    _no_wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    _no_wrap_left   = Alignment(horizontal="left", vertical="center", wrap_text=False, indent=1)
    _wrap_center    = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _merge_row(ws, 1, 1, total_cols,
               f"Позиция {position_idx}: {item_name}",
               fill_color="2E75B6",
               font=Font(bold=True, size=14, color="FFFFFF"),
               align=_wrap_center, height=32)

    if result.get("category_not_detected"):
        warn_text = "Категория не определена — подбор не выполнен"
        warn_color = "C00000"
    else:
        warn_text = "Проверьте модели по исходным таблицам каталога"
        warn_color = "FF8C00"
    _merge_row(ws, 2, 1, total_cols, warn_text,
               fill_color=warn_color,
               font=Font(bold=True, size=9, color="FFFFFF"),
               align=_no_wrap_center, height=20)

    # ── Строка 3: пустой разделитель ────────────────────────────────────────
    ws.row_dimensions[3].height = 4

    # Шапка: строки 4–6 (название модели, версия, процент)
    HDR_NAME = 4
    HDR_VER  = 5
    HDR_PCT  = 6

    # Левые колонки (№, Характеристика, Требуется) — merge по вертикали
    for ci, label, width in [
        (COL_NUM,  "№",              4),
        (COL_CHAR, "Характеристика", 40),
        (COL_REQ,  "Требуется",      20),
    ]:
        ws.merge_cells(start_row=HDR_NAME, start_column=ci,
                       end_row=HDR_PCT, end_column=ci)
        cell = ws.cell(row=HDR_NAME, column=ci, value=label)
        cell.font = _bold(11)
        cell.fill = _fill(COLOR_BLUE_HDR)
        cell.alignment = _wrap_center
        cell.border = _thin_border()
        for r in range(HDR_NAME, HDR_PCT + 1):
            ws.cell(row=r, column=ci).border = _thin_border()

    for ci_i, item in enumerate(top):
        ci = COL_FIRST + ci_i
        pct = item["pct"]
        bg = _pct_color(pct)

        c1 = ws.cell(row=HDR_NAME, column=ci, value=item["model_name"])
        c1.font = _bold(11)
        c1.fill = _fill(COLOR_BLUE_HDR)
        c1.alignment = _no_wrap_center
        c1.border = _thin_border()

        c2 = ws.cell(row=HDR_VER, column=ci, value=item["version"])
        c2.font = Font(size=9, color="555555")
        c2.fill = _fill(COLOR_BLUE_HDR)
        c2.alignment = _no_wrap_center
        c2.border = _thin_border()

        c3 = ws.cell(row=HDR_PCT, column=ci, value=f"{pct:.0f}%")
        c3.font = _bold(12)
        c3.fill = _fill(bg)
        c3.alignment = _no_wrap_center
        c3.border = _thin_border()

    ws.row_dimensions[HDR_NAME].height = 22
    ws.row_dimensions[HDR_VER].height = 16
    ws.row_dimensions[HDR_PCT].height = 22

    data_start = HDR_PCT + 1
    ZEBRA_A = "F7F9FC"
    ZEBRA_B = "FFFFFF"

    for ri, (key, req_val) in enumerate(required_specs.items()):
        row = data_start + ri
        readable = reverse_mapping.get(key, key.replace("_", " ").title())
        zebra = ZEBRA_A if ri % 2 == 0 else ZEBRA_B

        statuses = []
        for item in top:
            m = item["match_data"]
            if key in set(m.get("matched_specs", [])):
                statuses.append("match")
            elif key in set(m.get("unmapped_specs", m.get("missing_specs", []))):
                statuses.append("unmapped")
            elif key in m.get("different_specs", {}):
                statuses.append("diff")
            else:
                statuses.append("none")

        # № строки
        nc = ws.cell(row=row, column=COL_NUM, value=ri + 1)
        nc.font = Font(size=9, color="999999")
        nc.fill = _fill(zebra)
        nc.alignment = _no_wrap_center
        nc.border = _thin_border()

        cc = ws.cell(row=row, column=COL_CHAR, value=readable)
        cc.font = Font(size=10)
        cc.fill = _fill(zebra)
        cc.alignment = _no_wrap_left
        cc.border = _thin_border()

        rc = ws.cell(row=row, column=COL_REQ, value=_fmt_val(req_val))
        rc.font = Font(size=10, bold=True)
        rc.fill = _fill(zebra)
        rc.alignment = _no_wrap_center
        rc.border = _thin_border()

        for ci_i, (item, status) in enumerate(zip(top, statuses)):
            ci = COL_FIRST + ci_i
            m = item["match_data"]
            model_specs = m.get("attributes") or m.get("specifications") or {}
            mod_val = model_specs.get(key)

            if status == "match":
                cell_bg = COLOR_GREEN
                display = _fmt_val(mod_val)
                font = Font(size=10)
            elif status == "unmapped":
                cell_bg = COLOR_LIGHT_GRAY
                display = "—"
                font = Font(size=10, color="999999")
            elif status == "diff":
                cell_bg = COLOR_RED
                display = _fmt_val(mod_val)
                font = Font(size=10)
            else:
                cell_bg = zebra
                display = _fmt_val(mod_val) if mod_val is not None else "—"
                font = Font(size=10, color="666666")

            cell = ws.cell(row=row, column=ci, value=display)
            cell.fill = _fill(cell_bg)
            cell.font = font
            cell.alignment = _no_wrap_center
            cell.border = _thin_border()

        ws.row_dimensions[row].height = 22

    if top and required_specs:
        summary_row = data_start + len(required_specs)
        ws.row_dimensions[summary_row].height = 4

        totals_row = summary_row + 1
        ws.cell(row=totals_row, column=COL_CHAR,
                value="ИТОГО").font = _bold(11)
        ws.cell(row=totals_row, column=COL_REQ,
                value=f"{len(required_specs)} хар-к").font = _bold(10)
        ws.cell(row=totals_row, column=COL_REQ).alignment = _no_wrap_center
        for ci_i, item in enumerate(top):
            ci = COL_FIRST + ci_i
            bg = _pct_color(item["pct"])
            matched = item["matched"]
            total = item["total_specs"]
            cell = ws.cell(row=totals_row, column=ci,
                           value=f"{matched}/{total}")
            cell.font = _bold(11)
            cell.fill = _fill(bg)
            cell.alignment = _no_wrap_center
            cell.border = _thin_border()
        ws.row_dimensions[totals_row].height = 24

    ws.column_dimensions[get_column_letter(COL_NUM)].width = 5
    ws.column_dimensions[get_column_letter(COL_CHAR)].width = 42
    ws.column_dimensions[get_column_letter(COL_REQ)].width = 22
    for ci_i in range(n_models):
        # Ширина по длине названия модели (мин 18, макс 30)
        name_len = len(top[ci_i]["model_name"]) if ci_i < len(top) else 0
        w = max(18, min(name_len + 6, 30))
        ws.column_dimensions[get_column_letter(COL_FIRST + ci_i)].width = w

    # Закрепление шапки + левых колонок
    ws.freeze_panes = ws.cell(row=data_start, column=COL_FIRST)

    logger.info(f"Detail sheet '{sheet_name}': {len(required_specs)} specs, {n_models} models")


_OP_UNICODE = {">=": "≥", "<=": "≤", "!=": "≠", ">": ">", "<": "<", "=": "="}


def _fmt_op_str(s: str) -> str:
    """'>512' → '> 512',  '<=1024' → '≤ 1024'."""
    m = re.match(r'^([><=!]{1,2})\s*(.+)$', s.strip())
    if m:
        op = _OP_UNICODE.get(m.group(1), m.group(1))
        return f"{op} {m.group(2).strip()}"
    return s


def _fmt_val(val: Any) -> str:
    if val is None:
        return "—"
    if isinstance(val, bool):
        return "Да" if val else "Нет"
    if isinstance(val, list):
        return " и ".join(_fmt_op_str(str(v)) for v in val)
    if isinstance(val, str):
        return _fmt_op_str(val) if re.match(r'^[><=!]', val) else val
    return str(val)


def _effective_specs(specs: dict) -> dict:
    """Убрать внутренние ключи (например, __canonical__) из required_specs."""
    return {k: v for k, v in specs.items() if not k.startswith("__")}


# Устанавливается в generate_report из настроек приложения
_FILTER_BY_SPEC_COUNT: bool = True


def generate_report(
    requirements: Dict[str, Any],
    match_results: Dict[str, Any],
    output_dir: str = "temp_files",
    min_percentage: float = 75.0,
    filename: str = "",
    processing_time: float = 0.0,
) -> str:
    """
    Генерация Excel отчёта.

    Структура:
    - Лист "Сводка": предупреждение + метаданные + таблица по каждой позиции
    - Листы "Поз. N ...": детальное сравнение характеристик (одна позиция на лист)
    """
    global _FILTER_BY_SPEC_COUNT
    try:
        from config import settings
        # При EAV+fuzzy-матчинге 'specifications' содержит только совпавшие значения,
        # поэтому filter_by_spec_count может ложно исключать подходящие модели.
        _FILTER_BY_SPEC_COUNT = settings.filter_by_spec_count
    except Exception:
        _FILTER_BY_SPEC_COUNT = False

    logger.info(
        f"Генерация Excel-отчёта (зелёный≥{THRESHOLD_GREEN}%, жёлтый≥{THRESHOLD_YELLOW}%, "
        f"мин≥{THRESHOLD_MIN}%, filter_by_spec_count={_FILTER_BY_SPEC_COUNT})…"
    )

    wb = Workbook()
    results = match_results.get("results", [])

    _create_summary_sheet(wb, match_results, filename, processing_time)

    for idx, result in enumerate(results, 1):
        _create_detail_sheet(wb, result, idx, max_models=10)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(output_dir, f"tender_match_report_{timestamp}.xlsx")
    wb.save(file_path)

    logger.info(f"Excel сохранён: {file_path} ({len(wb.sheetnames)} листов)")
    return file_path
